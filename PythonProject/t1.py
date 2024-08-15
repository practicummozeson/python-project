import os
from flask import Flask, request, jsonify
import pandas as pd
from openpyxl.reader.excel import load_workbook
from fpdf import FPDF

app = Flask(__name__)
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'
    file.save(file.filename)
    file_url = os.path.abspath(file.filename)
    workbook = load_workbook(filename=file_url)
    number_of_sheets = len(workbook.sheetnames)
    return {"file_url": file_url,
            "number_of_sheets": number_of_sheets}
def generate_report(file_path, sheet_list):
    df = pd.read_excel(file_path)

    report = {}
    for sheet_data in sheet_list:
        sheet_name = sheet_data['sheet_name']
        action = sheet_data['action']
        columns = sheet_data['columns']

        if sheet_name in df.sheet_names:
            sheet_df = df.parse(sheet_name)
            selected_columns = sheet_df[columns]

            if action == 'option':
                report[sheet_name] = selected_columns.max().to_dict()
            elif action == 'average':
                report[sheet_name] = selected_columns.mean().to_dict()
            else:
                report[sheet_name] = "Invalid action specified"

        else:
            report[sheet_name] = "Sheet not found in the file"

    return report

def create_pdf_report(report_data, pdf_file_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    for sheet_name, sheet_info in report_data.items():
        pdf.cell(200, 10, txt=f"Sheet: {sheet_name}", ln=True)
        for col_name, value in sheet_info.items():
            pdf.cell(200, 10, txt=f"Column: {col_name}, Value: {value}", ln=True)

    pdf.output(pdf_file_path)

@app.route('/generate_report', methods=['POST'])
def handle_generate_report():
    data = request.json
    file_path = data.get('file_path')
    sheet_list = data.get('sheet_list')

    if not file_path or not sheet_list:
        return jsonify({'error': 'Missing file_path or sheet_list'}), 400

    report_data = generate_report(file_path, sheet_list)
    return jsonify(report_data)

@app.route('/generate_pdf_report', methods=['POST'])
def handle_generate_pdf_report():
    data = request.json
    report_data = data.get('report_data')

    if not report_data:
        return jsonify({'error': 'Missing report_data'}), 400

    pdf_file_path = 'report.pdf'
    create_pdf_report(report_data, pdf_file_path)

    return jsonify({'message': 'PDF report generated successfully'})
if __name__ == '__main__':
    app.run()